# coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from xlutils.copy import copy


class OperationExcel:
    def __init__(self, file_name=None, sheet_id=None, copy_file_name=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id
        else:
            self.file_name = '../dataConfig/interface.xlsx'
            self.sheet_id = 0
        if copy_file_name is None:
            self.data = self.get_data()
        else:
            # 如果是输出文件，则
            data = load_workbook(copy_file_name)
            data.copy_worksheet(data.worksheets[0])
            data.save(self.file_name)
            self.data = self.get_data()

    def get_all_data(self):
        data = load_workbook(self.file_name)
        return data

    def get_data(self):
        data = load_workbook(self.file_name)
        sheet_names = data.sheetnames
        table = data[sheet_names[self.sheet_id]]

        return table

    def get_lines(self):
        return self.data.max_row

    def get_cell_value(self, row, col):
        return self.data.cell(row, col).value

    # 写入数据
    def write_value(self, row, col, value):
        write_data = load_workbook(self.file_name)
        sheet_names = write_data.get_sheet_names()
        sheet_data = write_data.get_sheet_by_name(sheet_names[0])
        sheet_data.cell(row, col).value = value
        write_data.save(self.file_name)

    # 根据对应的case_id找到对应行的内容
    def get_rows_data(self, case_id):
        return self.get_row_values(self.get_row_num(case_id))

    # 根据对应的case_id找到对应的行号
    def get_row_num(self, case_id):
        num = 2
        cols_data = self.get_cols_data()
        print(cols_data)
        for col_data in cols_data:
            if case_id in col_data:
                return num
            num = num + 1

    # 根据行号，找到该行的内容
    def get_row_values(self, row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data

    # 获取某一列的内容
    def get_cols_data(self, col_id=None):
        cols = []
        if col_id is not None:
            for i in range(2, self.get_lines()+1):
                cols.append(self.get_cell_value(i, col_id))
        else:
            for i in range(2, self.get_lines()+1):
                cols.append(self.get_cell_value(i, 1))
        return cols


if __name__ == '__main__':
    # case文件
    data_file_name = '../dataConfig/interface.xlsx'
    # result文件
    result_file_name = '../dataConfig/out.xlsx'
    opers1 = OperationExcel(data_file_name, 0)
    opers2 = OperationExcel(result_file_name, 0, data_file_name)

    print(opers1.get_cell_value(2, 2))
